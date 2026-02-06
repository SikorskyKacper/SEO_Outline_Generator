import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_template(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arkusz1"

    # Setup Columns
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 80

    # Key headers
    ws['B2'] = "Słowo kluczowe"
    ws['C2'] = "Search Volume Average"
    header_font = Font(bold=True)
    ws['B2'].font = header_font
    ws['C2'].font = header_font

    # Metadata Placeholders
    labels = {
        "E3": "Meta title",
        "E4": "Meta description",
        "E5": "H1",
        "E6": "Lead"
    }
    for cell, val in labels.items():
        ws[cell] = val
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='right')

    # Example H2 (Row 7)
    ws['E7'] = "H2"
    ws['F7'] = "Przykładowy Nagłówek H2"
    ws['E7'].font = Font(bold=True, size=12)
    ws['F7'].font = Font(bold=True, size=12)
    ws['E7'].fill = PatternFill("solid", fgColor="DDDDDD")
    ws['F7'].fill = PatternFill("solid", fgColor="DDDDDD")

    # Example Brief (Row 8)
    ws['E8'] = "Treść do nagłówka"
    ws['F8'] = "- Punkt 1\n- Punkt 2\n- Punkt 3"
    ws['E8'].alignment = Alignment(vertical='top', horizontal='right')
    ws['F8'].alignment = Alignment(wrap_text=True)

    # Example H3 (Row 9)
    ws['E9'] = "H3"
    ws['F9'] = "Przykładowy Nagłówek H3"
    ws['E9'].font = Font(bold=True, italic=True)
    ws['F9'].font = Font(bold=True, italic=True)

    # Example Brief (Row 10)
    ws['E10'] = "Treść do nagłówka"
    ws['F10'] = "- Szczegół 1\n- Szczegół 2"
    ws['E10'].alignment = Alignment(vertical='top', horizontal='right')
    ws['F10'].alignment = Alignment(wrap_text=True)

    wb.save(path)
    print(f"Template created at {path}")

if __name__ == "__main__":
    os.makedirs("templates", exist_ok=True)
    create_template("templates/konspekt_template.xlsx")
