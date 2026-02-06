import openpyxl
from openpyxl.styles import Alignment, Font
from schemas import Outline, Keyword
from typing import List
import copy
import os

def write_to_excel(
    outline: Outline, 
    keywords: List[Keyword], 
    template_path: str, 
    output_path: str
):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Arkusz1"]

    # 1. Fill Keywords (B3:C17)
    # Ensure we strictly follow the 15 rows requirement
    for i in range(15):
        row_idx = 3 + i
        if i < len(keywords):
            ws[f"B{row_idx}"] = keywords[i].keyword
            ws[f"C{row_idx}"] = keywords[i].search_volume
        else:
            ws[f"B{row_idx}"] = ""
            ws[f"C{row_idx}"] = ""

    # 2. Fill Static Fields (F3:F6)
    ws["F3"] = outline.meta_title
    ws["F4"] = outline.meta_description
    ws["F5"] = outline.h1_title
    ws["F6"] = outline.lead

    # 3. Handle Outline (Row 7+)
    start_row = 7
    
    # 3.1 Clear existing content in structure columns
    # Find last row
    max_row = ws.max_row
    for r in range(start_row, max_row + 20): # Clear buffer
        ws[f"E{r}"] = None
        ws[f"F{r}"] = None

    # 3.2 Prepare styles for copying
    # We need to find ONE H2 header row and ONE Brief row to use as source
    # Assuming the template has at least one of each around 7-10
    # If not, we fall back to defaults, but the spec says use template styles.
    
    # Let's try to grab styles from E7/F7 (Header) and E8/F8 (Brief) as referenced in prompt
    header_style_source_cell_e = ws["E7"]
    header_style_source_cell_f = ws["F7"]
    brief_style_source_cell_e = ws["E8"]
    brief_style_source_cell_f = ws["F8"]

    current_row = start_row
    
    for item in outline.items:
        # --- Header Row ---
        # Apply style
        for col in ["E", "F"]:
            target = ws[f"{col}{current_row}"]
            source = ws[f"{col}7"] # Always copy from row 7 (Header Source)
            
            target.font = copy.copy(source.font)
            target.border = copy.copy(source.border)
            target.fill = copy.copy(source.fill)
            target.number_format = copy.copy(source.number_format)
            target.protection = copy.copy(source.protection)
            target.alignment = copy.copy(source.alignment)
            
        # Set values
        ws[f"E{current_row}"] = item.level 
        ws[f"F{current_row}"] = item.heading
        
        # Adjust row height if present in source
        ws.row_dimensions[current_row].height = ws.row_dimensions[7].height

        current_row += 1

        # --- Brief Row ---
        # Apply style
        for col in ["E", "F"]:
            target = ws[f"{col}{current_row}"]
            source = ws[f"{col}8"] # Always copy from row 8 (Brief Source)
            
            target.font = copy.copy(source.font)
            target.border = copy.copy(source.border)
            target.fill = copy.copy(source.fill)
            target.number_format = copy.copy(source.number_format)
            target.protection = copy.copy(source.protection)
            target.alignment = copy.copy(source.alignment)
            
        # Set values
        ws[f"E{current_row}"] = "Treść do nagłówka"
        brief_text = "\n".join([f"- {point}" for point in item.brief])
        ws[f"F{current_row}"] = brief_text
        
        # Ensure wrap text
        ws[f"F{current_row}"].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust height (rough estimation: 15px per line)
        line_count = len(item.brief) + 1
        ws.row_dimensions[current_row].height = max(15 * line_count, ws.row_dimensions[8].height or 15)

        current_row += 1

    wb.save(output_path)
